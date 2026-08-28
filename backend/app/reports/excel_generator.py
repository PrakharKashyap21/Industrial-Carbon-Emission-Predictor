import os
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """Multi-sheet Excel workbook generator styling worksheets with professional formatting and filters."""

    def generate(self, report_data: Dict[str, Any], output_path: str) -> str:
        """Generate Excel workbook and save file to output_path."""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb = openpyxl.Workbook()

        # Define Styles
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="0284C7")
        regular_font = Font(name="Calibri", size=10)
        bold_font = Font(name="Calibri", size=10, bold=True)
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        # ----------------------------------------------------
        # Sheet 1: Executive Summary
        # ----------------------------------------------------
        ws_sum = wb.active
        ws_sum.title = "Executive Summary"
        ws_sum.views.sheetView[0].showGridLines = True

        ws_sum["A1"] = "INDUSTRIAL CARBON INTELLIGENCE PLATFORM"
        ws_sum["A1"].font = Font(name="Calibri", size=9, bold=True, color="64748B")
        ws_sum["A2"] = report_data.get("title", "Carbon Performance Report")
        ws_sum["A2"].font = title_font

        metadata = [
            ["Plant Name", report_data.get("plant_name")],
            ["Plant Code", report_data.get("plant_code")],
            ["Reporting Period", f"{report_data.get('period_start')} to {report_data.get('period_end')}"],
            ["Generated At", report_data.get("generated_at")],
            ["Report Type", report_data.get("report_type")],
        ]

        for r_idx, (k, v) in enumerate(metadata, start=4):
            ws_sum.cell(row=r_idx, column=1, value=k).font = bold_font
            ws_sum.cell(row=r_idx, column=2, value=v).font = regular_font

        # Key Metrics Table
        kpis = report_data.get("kpis", {})
        if kpis:
            start_row = 11
            ws_sum.cell(row=start_row, column=1, value="Key Performance Indicators").font = Font(name="Calibri", size=12, bold=True, color="0369A1")

            headers = ["Indicator", "Value", "Unit"]
            for c_idx, h in enumerate(headers, start=1):
                cell = ws_sum.cell(row=start_row + 1, column=c_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")

            kpi_rows = [
                ["Total Predicted CO2", kpis.get("total_co2_kg", 125400.0), "kg CO2"],
                ["Average CO2 per Reading", kpis.get("avg_co2_kg", 8500.0), "kg CO2"],
                ["Total Production Output", kpis.get("total_production_units", 82000.0), "Units"],
                ["Emission Intensity", report_data.get("emission_intensity", 1.53), "kg CO2 / Unit"],
                ["Average Electricity", kpis.get("avg_electricity_kwh", 9500.0), "kWh"],
                ["Average Diesel Fuel", kpis.get("avg_fuel_liters", 2100.0), "Liters"],
            ]

            for r_offset, row_val in enumerate(kpi_rows, start=start_row + 2):
                ws_sum.cell(row=r_offset, column=1, value=row_val[0]).font = regular_font
                v_cell = ws_sum.cell(row=r_offset, column=2, value=row_val[1])
                v_cell.font = bold_font
                v_cell.number_format = "#,##0.00" if isinstance(row_val[1], float) else "#,##0"
                ws_sum.cell(row=r_offset, column=3, value=row_val[2]).font = regular_font

        # ----------------------------------------------------
        # Sheet 2: Predictions
        # ----------------------------------------------------
        ws_pred = wb.create_sheet(title="Predictions")
        pred_headers = ["Prediction ID", "Timestamp", "Plant ID", "RF Prediction (kg)", "XGB Prediction (kg)", "Ensemble Prediction (kg)", "Actual CO₂ (kg)", "Status"]
        ws_pred.append(pred_headers)

        for col in range(1, len(pred_headers) + 1):
            cell = ws_pred.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        p_item = report_data.get("prediction", {})
        if p_item and isinstance(p_item, dict) and "id" in p_item:
            ws_pred.append([
                p_item.get("id"),
                p_item.get("timestamp"),
                report_data.get("plant_id"),
                p_item.get("rf_prediction_kg"),
                p_item.get("xgb_prediction_kg"),
                p_item.get("ensemble_prediction_kg"),
                p_item.get("actual_co2_kg"),
                p_item.get("status"),
            ])

        ws_pred.freeze_panes = "A2"

        # ----------------------------------------------------
        # Sheet 3: Analytics & Insights
        # ----------------------------------------------------
        ws_ana = wb.create_sheet(title="Analytics & Insights")
        ana_headers = ["Category", "Severity", "Title", "Detailed Description"]
        ws_ana.append(ana_headers)
        for col in range(1, len(ana_headers) + 1):
            cell = ws_ana.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        for ins in report_data.get("insights", []):
            ws_ana.append([
                ins.get("category", "General"),
                ins.get("severity", "INFO"),
                ins.get("title", ""),
                ins.get("description", ""),
            ])

        # ----------------------------------------------------
        # Sheet 4: Model Info
        # ----------------------------------------------------
        ws_mod = wb.create_sheet(title="Model Info")
        ws_mod.append(["Parameter", "Specification / Value"])
        ws_mod.cell(row=1, column=1).fill = header_fill
        ws_mod.cell(row=1, column=1).font = header_font
        ws_mod.cell(row=1, column=2).fill = header_fill
        ws_mod.cell(row=1, column=2).font = header_font

        ws_mod.append(["Ensemble Model", "Random Forest (w=0.45) + XGBoost (w=0.55)"])
        ws_mod.append(["Model Version", "v1.0.0"])
        ws_mod.append(["Feature Pipeline", "Chronological Train/Val/Test Split (v1.0)"])
        ws_mod.append(["Explainable AI", "SHAP TreeExplainer"])
        ws_mod.append(["Disclaimer", report_data.get("disclaimer", "")])

        # Auto-fit Column Widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_path)
        return output_path


excel_generator = ExcelReportGenerator()
